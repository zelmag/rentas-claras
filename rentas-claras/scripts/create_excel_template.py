"""
RentasClaras: Excel Template Generator
======================================

Creates an Excel workbook with properly formatted tables for Microsoft Graph API.
The tables will be "clickable" and compatible with the excel_client.py API calls.

Run: python scripts/create_excel_template.py
Output: Inquilinos.xlsx (upload this to OneDrive/SharePoint)
"""

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


def create_rentasclaras_template(output_path: str = "Inquilinos.xlsx"):
    """Create Excel workbook with properly named tables."""
    
    wb = Workbook()
    
    # ==========================================================================
    # SHEET 1: Inquilinos (Tenants)
    # ==========================================================================
    ws_inquilinos = wb.active
    ws_inquilinos.title = "Inquilinos"
    
    # Headers
    inquilinos_headers = [
        "ID_Inquilino", "Nombre", "WhatsApp", "Propiedad", "Unidad",
        "Renta_Mensual", "Fecha_Ingreso", "Fecha_Salida", "Email"
    ]
    ws_inquilinos.append(inquilinos_headers)
    
    # Sample data (you can delete these rows later)
    sample_tenants = [
        ["T-001", "María González", "+528112345001", "Ensenada", "3", 4500, "2024-01-15", "", "maria@email.com"],
        ["T-002", "Carlos Mendoza", "+528112345002", "Ensenada", "5", 5200, "2024-06-01", "", "carlos@email.com"],
        ["T-003", "Ana Ramírez", "+528112345003", "Huichapan", "1", 3800, "2024-03-01", "", ""],
    ]
    for row in sample_tenants:
        ws_inquilinos.append(row)
    
    # Set column widths
    col_widths = [15, 25, 18, 15, 10, 15, 15, 15, 25]
    for i, width in enumerate(col_widths, 1):
        ws_inquilinos.column_dimensions[get_column_letter(i)].width = width
    
    # Create TABLE - THIS IS THE KEY PART!
    table_inquilinos = Table(
        displayName="TablaInquilinos",  # This name is used by Graph API
        ref=f"A1:I{len(sample_tenants) + 1}"
    )
    table_inquilinos.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws_inquilinos.add_table(table_inquilinos)
    
    # ==========================================================================
    # SHEET 2: Pagos (Payments)
    # ==========================================================================
    ws_pagos = wb.create_sheet("Pagos")
    
    pagos_headers = [
        "ID_Pago", "ID_Inquilino", "Fecha_Pago", "Monto", "Metodo",
        "Codigo_Retiro", "Banco", "Concepto", "Folio", "Confirmado", "Notas"
    ]
    ws_pagos.append(pagos_headers)
    
    # Sample data
    sample_payments = [
        ["P-2025-01-001", "T-001", "2025-01-05", 4850, "Retiro sin tarjeta", 
         "847293102934", "BBVA", "Renta Enero + Luz", "RC-20250105-00001", True, ""],
    ]
    for row in sample_payments:
        ws_pagos.append(row)
    
    # Column widths
    col_widths = [18, 15, 15, 12, 20, 18, 10, 25, 22, 12, 30]
    for i, width in enumerate(col_widths, 1):
        ws_pagos.column_dimensions[get_column_letter(i)].width = width
    
    # Create TABLE
    table_pagos = Table(
        displayName="TablaPagos",  # Used by Graph API in add_payment()
        ref=f"A1:K{len(sample_payments) + 1}"
    )
    table_pagos.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws_pagos.add_table(table_pagos)
    
    # ==========================================================================
    # Save workbook
    # ==========================================================================
    wb.save(output_path)
    print(f"✅ Excel template created: {output_path}")
    print()
    print("📋 Tables created:")
    print("   • TablaInquilinos (sheet: Inquilinos)")
    print("   • TablaPagos (sheet: Pagos)")
    print()
    print("📤 Next steps:")
    print("   1. Open the file and verify tables are clickable")
    print("   2. Upload to OneDrive: /RentasClaras/Inquilinos.xlsx")
    print("   3. Update your tenants data in the Inquilinos sheet")
    print("   4. Delete sample data rows (keep headers!)")
    
    return output_path


if __name__ == "__main__":
    # Create in the project root
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    output_file = os.path.join(project_root, "Inquilinos.xlsx")
    
    create_rentasclaras_template(output_file)
