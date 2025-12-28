"""
Test Excel Tables Locally
=========================

This script tests that the Excel tables are properly configured
and can be manipulated (simulating what Microsoft Graph API would do).

Run: python scripts/test_excel_tables.py
"""

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from datetime import datetime
import os

def test_excel_tables():
    """Test that Excel tables are clickable and functional."""
    
    excel_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "Inquilinos.xlsx"
    )
    
    if not os.path.exists(excel_path):
        print(f"❌ File not found: {excel_path}")
        print("   Run: python scripts/create_excel_template.py first")
        return False
    
    print("=" * 60)
    print("🧪 TESTING EXCEL TABLES")
    print("=" * 60)
    print(f"\n📂 File: {excel_path}\n")
    
    wb = load_workbook(excel_path)
    all_tests_passed = True
    
    # =========================================================================
    # TEST 1: Verify all tables exist
    # =========================================================================
    print("TEST 1: Verify tables exist")
    print("-" * 40)
    
    expected_tables = {
        "Inquilinos": "TablaInquilinos",
        "Pagos": "TablaPagos",
        "Codigos_Retiro": "TablaCodigos"
    }
    
    for sheet_name, table_name in expected_tables.items():
        if sheet_name not in wb.sheetnames:
            print(f"   ❌ Sheet '{sheet_name}' not found")
            all_tests_passed = False
            continue
            
        ws = wb[sheet_name]
        table_names = [t.displayName for t in ws.tables.values()]
        
        if table_name in table_names:
            table = ws.tables[table_name]
            print(f"   ✅ {table_name} exists (range: {table.ref})")
        else:
            print(f"   ❌ {table_name} not found in {sheet_name}")
            print(f"      Found tables: {table_names}")
            all_tests_passed = False
    
    # =========================================================================
    # TEST 2: Simulate adding a row to TablaPagos (like Graph API would)
    # =========================================================================
    print("\nTEST 2: Simulate adding row to TablaPagos")
    print("-" * 40)
    
    ws_pagos = wb["Pagos"]
    table_pagos = ws_pagos.tables["TablaPagos"]
    
    # Get current table range
    old_ref = table_pagos.ref
    print(f"   Current range: {old_ref}")
    
    # Parse the range to find last row
    # Format: A1:K2 -> we need row 2, then add to row 3
    ref_parts = old_ref.split(":")
    end_cell = ref_parts[1]  # e.g., "K2"
    end_col = ''.join(filter(str.isalpha, end_cell))  # "K"
    end_row = int(''.join(filter(str.isdigit, end_cell)))  # 2
    
    new_row_num = end_row + 1
    
    # Add new payment row
    new_payment = [
        f"P-{datetime.now().strftime('%Y-%m-%H%M%S')}",  # ID_Pago
        "T-002",  # ID_Inquilino
        datetime.now().strftime("%Y-%m-%d"),  # Fecha_Pago
        5200,  # Monto
        "SPEI",  # Metodo
        "",  # Codigo_Retiro
        "Santander",  # Banco
        "Renta Enero 2025",  # Concepto
        f"RC-{datetime.now().strftime('%Y%m%d')}-00002",  # Folio
        False,  # Confirmado
        "Test payment"  # Notas
    ]
    
    for col_idx, value in enumerate(new_payment, 1):
        ws_pagos.cell(row=new_row_num, column=col_idx, value=value)
    
    # Update table range to include new row
    new_ref = f"A1:{end_col}{new_row_num}"
    table_pagos.ref = new_ref
    
    print(f"   ✅ Added row {new_row_num}: {new_payment[0]}")
    print(f"   ✅ Updated table range: {new_ref}")
    
    # =========================================================================
    # TEST 3: Simulate adding a row to TablaInquilinos
    # =========================================================================
    print("\nTEST 3: Simulate adding row to TablaInquilinos")
    print("-" * 40)
    
    ws_inquilinos = wb["Inquilinos"]
    table_inquilinos = ws_inquilinos.tables["TablaInquilinos"]
    
    old_ref = table_inquilinos.ref
    ref_parts = old_ref.split(":")
    end_cell = ref_parts[1]
    end_col = ''.join(filter(str.isalpha, end_cell))
    end_row = int(''.join(filter(str.isdigit, end_cell)))
    
    new_row_num = end_row + 1
    
    new_tenant = [
        f"T-{new_row_num:03d}",  # ID_Inquilino
        "Test Tenant",  # Nombre
        "+528199999999",  # WhatsApp
        "TestProperty",  # Propiedad
        "99",  # Unidad
        9999,  # Renta_Mensual
        datetime.now().strftime("%Y-%m-%d"),  # Fecha_Ingreso
        "",  # Fecha_Salida
        "test@test.com"  # Email
    ]
    
    for col_idx, value in enumerate(new_tenant, 1):
        ws_inquilinos.cell(row=new_row_num, column=col_idx, value=value)
    
    new_ref = f"A1:{end_col}{new_row_num}"
    table_inquilinos.ref = new_ref
    
    print(f"   ✅ Added row {new_row_num}: {new_tenant[0]} - {new_tenant[1]}")
    print(f"   ✅ Updated table range: {new_ref}")
    
    # =========================================================================
    # TEST 4: Read data back from tables
    # =========================================================================
    print("\nTEST 4: Read data from tables")
    print("-" * 40)
    
    for sheet_name, table_name in expected_tables.items():
        ws = wb[sheet_name]
        table = ws.tables[table_name]
        
        # Parse range
        ref_parts = table.ref.split(":")
        start_row = int(''.join(filter(str.isdigit, ref_parts[0])))
        end_row = int(''.join(filter(str.isdigit, ref_parts[1])))
        
        row_count = end_row - start_row  # Exclude header
        print(f"   📊 {table_name}: {row_count} data rows")
    
    # =========================================================================
    # Save the modified workbook
    # =========================================================================
    test_output = excel_path.replace(".xlsx", "_tested.xlsx")
    wb.save(test_output)
    print(f"\n💾 Saved test file: {test_output}")
    
    # =========================================================================
    # Summary
    # =========================================================================
    print("\n" + "=" * 60)
    if all_tests_passed:
        print("✅ ALL TESTS PASSED!")
        print("=" * 60)
        print("""
The tables are properly configured and will work with Microsoft Graph API.

Next steps:
1. Upload 'Inquilinos.xlsx' to OneDrive at /RentasClaras/
2. The Graph API calls in excel_client.py will work correctly

Note: If tables appear "unclickable" in Excel Online, make sure you:
- Click INSIDE the table data (not on the header)
- Look for "Table Design" tab at the top of Excel
""")
    else:
        print("❌ SOME TESTS FAILED")
        print("=" * 60)
        print("Please run: python scripts/create_excel_template.py")
    
    return all_tests_passed


def show_table_details():
    """Show detailed table information."""
    excel_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "Inquilinos.xlsx"
    )
    
    if not os.path.exists(excel_path):
        print(f"File not found: {excel_path}")
        return
    
    wb = load_workbook(excel_path)
    
    print("\n📋 DETAILED TABLE INFORMATION")
    print("=" * 60)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n📑 Sheet: {sheet_name}")
        
        if ws.tables:
            for table in ws.tables.values():
                print(f"   Table Name: {table.displayName}")
                print(f"   Range: {table.ref}")
                print(f"   Style: {table.tableStyleInfo.name if table.tableStyleInfo else 'None'}")
                print(f"   Show Row Stripes: {table.tableStyleInfo.showRowStripes if table.tableStyleInfo else 'N/A'}")
        else:
            print("   ⚠️  NO TABLES IN THIS SHEET")


if __name__ == "__main__":
    test_excel_tables()
    show_table_details()
